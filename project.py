import openpyxl
import os
from PyPDF2 import PdfFileMerger, PdfFileReader
from fpdf import FPDF
import inflect
import datetime

def main():
    create_pdf()
    merge_pdfs()


def create_pdf():
    """ Collects the company details such as:
        Company Name, Address, Contact Number, and Email
        Loads the workbook that contains employee data and
        Identifies the worksheet to workwith

        Creates a PDF file for each iteration of the for loop
        Creates a directory 'output' to save all individual PDF's
    """
    
    # Input for company details
    comp_name = input("Company Name: ")
    comp_add = input("Company Address: ")
    comp_contact = input("Contact Number: ")
    comp_email = input("Company Email: ")
    month = input("Month and Year: ")

    # Loads the workbook that contains the employee data
    wb = openpyxl.load_workbook("D:\\Python Projects\\project\\emp_salary_details.xlsx") 
    sheet = wb['payroll']
    
    # Iterage the data inside the excel file per row of employee ang gets the data into variables
    # sheet.max_row gets the number of rows with data, incremented by 1 to get the last row
    for i in range(2, sheet.max_row+1):
        try:
            # Enters the data for each row into the variables
            emp_id = sheet.cell(row = i, column = 1).value
            emp_name = sheet.cell(row = i, column = 2).value
            basic_salary = sheet.cell(row = i, column = 3).value
            allowance = sheet.cell(row = i, column = 4).value
            days_worked = sheet.cell(row = i, column = 5).value
            ot_hours = sheet.cell(row = i, column = 6).value
            abs_days = sheet.cell(row = i, column = 7).value
            insurance = sheet.cell(row = i, column = 8).value
            other_deds = sheet.cell(row = i, column = 9).value
            ot_pay = (int(basic_salary) + int(allowance)) / 30 / 8 * 1.5 * int(ot_hours)
            ot_pay = int(ot_pay)
            total_earnings = int(basic_salary) + int(allowance) + int(ot_pay)
            abs_amount = (int(basic_salary) + int(allowance)) / 30 * int(abs_days)
            total_deds = int(abs_amount) + int(insurance) + int(other_deds)
            net_salary = total_earnings - total_deds

            # Sets the PDF format and inserts the data accordingly
            pdf = FPDF("P", "mm", "A4")
            w = 210
            h = 297
            pdf.set_left_margin(20)
            pdf.set_top_margin(20)
            pdf.set_right_margin(20)

            pdf.add_page()
            # Company logo goes here
            pdf.image("imgplaceholder.jpg", x = 20, y = 20, w = 50)

            # Set font and position for company name
            pdf.set_font("Helvetica", "BU", 22)
            pdf.set_y(30)
            pdf.set_x(75)
            pdf.cell(0, 15, comp_name, 0, align = "C")

            # Set font and position for company address
            pdf.set_font("Helvetica", "I", 12)
            pdf.set_y(40)
            pdf.set_x(75)
            pdf.cell(0, 15, txt = f"Address: {comp_add}", border = 0, align = "C")

            # Set font and position for company contact number
            pdf.set_y(45)
            pdf.set_x(75)
            pdf.cell(0, 15, txt = f"Contact #: {comp_contact}", border = 0, align = "C")

            # Set font and position for company contact email
            pdf.set_y(50)
            pdf.set_x(75)
            pdf.cell(0, 15, txt = f"Email: {comp_email}", border = 0, align = "C")

            # Prints a line
            pdf.set_y(70)
            pdf.cell(0, 0.5, "", border = "T", align = "C")

            # Prints Title Heading "Salary Slip for ..."
            pdf.set_font("Helvetica", "BU", 18)
            pdf.set_y(75)
            pdf.cell(0, 15, txt = f"Payslip for the month of {month}", border = 1, align = "C")

            # Prints "Employee Details"
            pdf.set_font("Helvetica", "B", 14)
            pdf.set_y(100)
            pdf.cell(95, 10, txt = "Employee Details", border = "", align = "L")

            # Employee ID
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_y(110)
            pdf.set_x(25)
            pdf.cell(25, 5, txt = "Emp ID: ", border = "LB", align = "L")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_x(50)
            pdf.cell(45, 5, txt = f"{emp_id}", border = "B", align = "L")
            # Employee Name
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_y(115)
            pdf.set_x(25)
            pdf.cell(25, 5, txt = "Emp Name: ", border = "LB", align = "L")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_x(50)
            pdf.cell(45, 5, txt = f"{emp_name}", border = "B", align = "L")
            # Basic Salary
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_y(120)
            pdf.set_x(25)
            pdf.cell(25, 5, txt = "Basic Salary:", border = "LB", align = "L")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_x(50)
            pdf.cell(45, 5, txt = f"{basic_salary}", border = "B", align = "R")
            # Allowances
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_y(125)
            pdf.set_x(25)
            pdf.cell(25, 5, txt = "Allowance:", border = "LB", align = "L")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_x(50)
            pdf.cell(45, 5, txt = f"{allowance}", border = "B", align = "R")
            # No. of Days Worked
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_y(115)
            pdf.set_x(105)
            pdf.cell(45, 5, txt = "No. of Days Worked: ", border = "LB", align = "L")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_x(150)
            pdf.cell(25, 5, txt = f"{days_worked}", border = "B", align = "R")
            # No. of OT Hours
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_y(120)
            pdf.set_x(105)
            pdf.cell(45, 5, txt = "No. of OT Hours: ", border = "LB", align = "L")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_x(150)
            pdf.cell(25, 5, txt = f"{ot_hours}", border = "B", align = "R")
            # No. of Absences
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_y(125)
            pdf.set_x(105)
            pdf.cell(45, 5, txt = "No. of Absent Days ", border = "LB", align = "L")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_x(150)
            pdf.cell(25, 5, txt = f"{abs_days}", border = "B", align = "R")

            # Salary Calculation Details
            pdf.set_font("Helvetica", "B", 14)
            pdf.set_y(140)
            pdf.cell(95, 10, txt = "Salary Calculation Details", border = "", align = "L")

            # EARNINGS
            # Basic Salary for 30 days
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_y(150)
            pdf.set_x(25)
            pdf.cell(25, 5, txt = "Basic Salary: ", border = "LB", align = "L")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_x(50)
            pdf.cell(45, 5, txt = f"{basic_salary}", border = "B", align = "R")
            # Allowance for 30 days
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_y(155)
            pdf.set_x(25)
            pdf.cell(25, 5, txt = "Allowance: ", border = "LB", align = "L")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_x(50)
            pdf.cell(45, 5, txt = f"{allowance}", border = "B", align = "R")
            # Overtime = Basic + Allowance / 30 / 8 * 1.5 * # of OT Hours
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_y(160)
            pdf.set_x(25)
            pdf.cell(25, 5, txt = "Overtime Pay: ", border = "LB", align = "L")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_x(50)
            pdf.cell(45, 5, txt = f"{ot_pay}", border = "B", align = "R")
            # Total Earnings = Basic + Allowance + OT
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_y(165)
            pdf.set_x(25)
            pdf.cell(25, 5, txt = "TOTAL EARNINGS: ", border = "LB", align = "L")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_x(50)
            pdf.cell(45, 5, txt = f"{total_earnings}", border = "B", align = "R")

            # DEDUCTIONS
            # Absences = Basic + Allowance / 30 * # of abs_days
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_y(150)
            pdf.set_x(105)
            pdf.cell(45, 5, txt = "Absences: ", border = "LB", align = "L")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_x(150)
            pdf.cell(25, 5, txt = f"{abs_amount}", border = "B", align = "R")
            # Insurance
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_y(155)
            pdf.set_x(105)
            pdf.cell(45, 5, txt = "Insurance: ", border = "LB", align = "L")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_x(150)
            pdf.cell(25, 5, txt = f"{insurance}", border = "B", align = "R")
            # Other Deductions
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_y(160)
            pdf.set_x(105)
            pdf.cell(45, 5, txt = "Other Deductions: ", border = "LB", align = "L")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_x(150)
            pdf.cell(25, 5, txt = f"{other_deds}", border = "B", align = "R")
            # TOTAL DEDUCTIONS
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_y(165)
            pdf.set_x(105)
            pdf.cell(45, 5, txt = "TOTAL DEDUCTIONS: ", border = "LB", align = "L")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_x(150)
            pdf.cell(25, 5, txt = f"{total_deds}", border = "B", align = "R")

            # NET SALARY = Total Earnings - Total Deductions
            pdf.set_font("Helvetica", "B", 14)
            pdf.set_y(180)
            pdf.cell(85, 10, txt = "NET SALARY", border = 1, align = "C")
            pdf.set_x(105)
            pdf.cell(85, 10, txt = f"{net_salary}", border = 1, align = "C")
            # in words
            net_in_words = amount_in_words(net_salary)
            pdf.set_font("Helvetica", "", 10)
            pdf.set_y(190)
            pdf.cell(0, 10, txt = f"AMOUNT IN WORDS : {net_in_words} SAUDI RIYALS" , border = 0, align = "L")


            pdf.set_font("Helvetica", "", 8)
            pdf.set_y(200)
            pdf.cell(0, 5, txt = "*****This is a computer-generated pay slip, signature not required.*****", border = 0, align = "C")

            printdate = get_print_date().strftime("%d/%m/%Y %H:%M:%S")
            pdf.set_font("Helvetica", "I", 8)
            pdf.set_y(230)
            pdf.cell(0, 5, txt = f"{printdate}", border = 0, align = "R")

            output_path = "D:\\Python Projects\\project\\output\\"
            if not os.path.exists(output_path):
                os.makedirs(output_path)
            pdf.output(output_path + str(emp_id) + ".pdf")
        except ValueError:
            pass
        
def amount_in_words(net_salary):
    """ Converts the int value of net salary into words
        Returs the string in uppercases """
    p = inflect.engine()
    return p.number_to_words(net_salary).upper()

def get_print_date():
    """ Gets and returns the current system datetime """
    return datetime.datetime.now()
    
def merge_pdfs():

    """ Gets the location of the individual pdf files
        and selects all files that ends with .pdf
        Creates an empty tuple and appends it for every .pdf file found
        Checks for the directory of merged pdfs exists or not,
        Creates a directory for merged PDFs
        Merges all the PDFs, saves it, and closes the funtion."""

    files_dir = "D:\\Python Projects\\project\\output"
    pdf_files = [f for f in os.listdir(files_dir) if f.endswith('.pdf')]
    merger = PdfFileMerger()
    for filename in pdf_files:
        merger.append(PdfFileReader(os.path.join(files_dir,filename),'rb'))
    merged_dir = "D:\\Python Projects\\project\\merged_dir"
    if not os.path.exists(merged_dir):
        os.makedirs(merged_dir)
    merger.write(os.path.join(merged_dir,'merged_pdfs.pdf'))
    merger.close()

if __name__ == "__main__":
    main()